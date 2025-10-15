"""
Export utilities for CSV, Excel, and PDF generation.

Provides functions to export data in various formats for download.
"""

import csv
import io
from datetime import datetime
from typing import Any, Dict, List

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ============================================================================
# CSV Export
# ============================================================================


def export_to_csv(
    data: List[Dict[str, Any]],
    filename: str,
    columns: List[str] = None
) -> StreamingResponse:
    """
    Export data to CSV format.

    Args:
        data: List of dictionaries containing the data
        filename: Name of the file to download
        columns: List of column names to include (None = all columns)

    Returns:
        StreamingResponse with CSV data
    """
    if not data:
        # Return empty CSV with headers only
        output = io.StringIO()
        if columns:
            writer = csv.DictWriter(output, fieldnames=columns)
            writer.writeheader()

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    # Determine columns
    if not columns:
        columns = list(data[0].keys())

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()

    for row in data:
        # Filter to include only specified columns
        filtered_row = {k: v for k, v in row.items() if k in columns}
        writer.writerow(filtered_row)

    # Return as streaming response
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ============================================================================
# Excel Export
# ============================================================================


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "Data",
    columns: List[str] = None,
    column_widths: Dict[str, int] = None
) -> StreamingResponse:
    """
    Export data to Excel (xlsx) format.

    Args:
        data: List of dictionaries containing the data
        filename: Name of the file to download
        sheet_name: Name of the worksheet
        columns: List of column names to include (None = all columns)
        column_widths: Dictionary mapping column names to widths

    Returns:
        StreamingResponse with Excel data
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        # Return empty workbook with headers only
        if columns:
            ws.append(columns)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    # Determine columns
    if not columns:
        columns = list(data[0].keys())

    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data.get(col_name, "")

            # Format datetime objects
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set column widths
    if column_widths:
        for col_name, width in column_widths.items():
            if col_name in columns:
                col_idx = columns.index(col_name) + 1
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = len(col_name)
            for row_data in data:
                value = str(row_data.get(col_name, ""))
                max_length = max(max_length, len(value))

            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ============================================================================
# PDF Export
# ============================================================================


def export_to_pdf(
    title: str,
    data: List[Dict[str, Any]],
    filename: str,
    columns: List[str] = None,
    column_labels: Dict[str, str] = None,
    page_size=A4,
    orientation="portrait"
) -> StreamingResponse:
    """
    Export data to PDF format as a table.

    Args:
        title: Document title
        data: List of dictionaries containing the data
        filename: Name of the file to download
        columns: List of column names to include (None = all columns)
        column_labels: Dictionary mapping column names to display labels
        page_size: Page size (A4 or letter)
        orientation: Page orientation ("portrait" or "landscape")

    Returns:
        StreamingResponse with PDF data
    """
    # Create PDF in memory
    output = io.BytesIO()

    # Adjust page size for landscape
    if orientation == "landscape":
        page_size = (page_size[1], page_size[0])  # Swap width and height

    # Create document
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )

    # Build story (content)
    story = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor("#366092"),
        spaceAfter=20,
        alignment=1  # Center
    )
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Timestamp
    timestamp_text = f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    timestamp_style = ParagraphStyle(
        'Timestamp',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=2  # Right
    )
    story.append(Paragraph(timestamp_text, timestamp_style))
    story.append(Spacer(1, 0.3 * inch))

    if not data:
        # No data message
        story.append(Paragraph("No data available.", styles['Normal']))
    else:
        # Determine columns
        if not columns:
            columns = list(data[0].keys())

        # Prepare table data
        table_data = []

        # Header row
        header_row = []
        for col in columns:
            label = column_labels.get(col, col) if column_labels else col
            header_row.append(label)
        table_data.append(header_row)

        # Data rows
        for row_data in data:
            row = []
            for col in columns:
                value = row_data.get(col, "")

                # Format values
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                elif isinstance(value, float):
                    value = f"{value:.2f}"
                elif value is None:
                    value = ""

                row.append(str(value))
            table_data.append(row)

        # Create table
        # Calculate column widths based on page size
        available_width = page_size[0] - inch  # 1 inch total margins
        col_width = available_width / len(columns)

        table = Table(table_data, colWidths=[col_width] * len(columns))

        # Style table
        table.setStyle(TableStyle([
            # Header style
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#366092")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows style
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        story.append(table)

        # Summary footer
        story.append(Spacer(1, 0.3 * inch))
        summary_text = f"Total records: {len(data)}"
        story.append(Paragraph(summary_text, styles['Normal']))

    # Build PDF
    doc.build(story)

    # Return as streaming response
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ============================================================================
# Helper Functions
# ============================================================================


def generate_filename(prefix: str, extension: str, include_timestamp: bool = True) -> str:
    """
    Generate a filename for export.

    Args:
        prefix: Filename prefix (e.g., "transactions", "users")
        extension: File extension without dot (e.g., "csv", "xlsx", "pdf")
        include_timestamp: Whether to include timestamp in filename

    Returns:
        Formatted filename
    """
    if include_timestamp:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{extension}"
    else:
        return f"{prefix}.{extension}"
